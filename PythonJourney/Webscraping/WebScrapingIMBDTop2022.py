from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill

linkList = [
    'https://www.imdb.com/search/title/?year=2022&title_type=feature,tv_series&explore=countries']


def get_IMDB_movie_list(link_list):
    for link in link_list:
        try:
            source = requests.get(link)
            source.raise_for_status()

            soup = BeautifulSoup(source.text, 'html.parser')

            movies = soup.find(
                'div', class_='lister-list').find_all('div', class_='lister-item mode-advanced')
            print(len(movies))

            excel = Workbook()
            sheet = excel.active
            sheet.title = "Top Rated IMDB 2022"
            # font = Font(bold=True)
            # a1 = sheet["'Name', 'Genre', 'Rating'"]
            # a1.ft = font
            # sheet.append(['Name', 'Genre', 'Rating'])
            sheet.append(['Name', 'Genre', 'Rating'])
            #top_row=['A1', 'B1', 'C1']
            # Bold first row
            font = Font(bold=True)
            for cell in sheet["1:1"]:
                cell.font = font        
            # sheet.append([ 'Name', 'Genre', 'Rating'])

            for movie in movies:
                number = movie.find(
                    'span', class_='lister-item-index unbold text-primary').text.strip()
                name = movie.find(
                    'h3', class_='lister-item-header').a.text.strip()
                genre = movie.find(
                    'div', class_='lister-item-content').find('span', class_='genre').text.strip()
                rating = movie.find(
                    'div', class_='lister-item-content').find('div', class_='inline-block ratings-imdb-rating').text.strip()

                print(number + name + ' ' + genre + ' ' + rating)

                sheet.append([name, genre, rating])
            
            thin_border = Border(top= Side(style='thin'), bottom= Side(style='thin'))
            # Insert top and bottom orders in each cell
            # range_string = "1:"+str(len(movies))
            # print(range_string)
            count = 0
            for cell in sheet["2:"+str(len(movies)+1)]:
                for singlecell in cell:
                    if count%2 == 0:
                        singlecell.fill = PatternFill(
                            "solid", start_color="AEAEAE")
                    else:
                        pass
                    print(singlecell)
                    singlecell.border = thin_border
                count+=1

            excel.save("Top Rated IMDB 2022.xlsx")

        except Exception as e:
            print(e)
    return None


get_IMDB_movie_list(linkList)
