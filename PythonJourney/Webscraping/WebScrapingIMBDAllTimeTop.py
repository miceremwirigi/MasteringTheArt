from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill

linkList = [
    'https://www.imdb.com/chart/toptv/']


def get_IMDB_movie_list(link_list):
    for link in link_list:
        try:
            source = requests.get(link)
            source.raise_for_status()

            #print(source.text)

            soup = BeautifulSoup(source.text, 'html.parser')

            
            # return a list of tr showing each movie details                
            movie_hrefs = soup.find(
                'div', class_='lister').find('tbody', class_='lister-list'). find_all('tr')
            for movie_href in movie_hrefs:
                #movies= []
                #movie = BeautifulSoup(movie_href.text, 'html.parser')
                # print(movie_href)
                
                #print(movie)
                #movies.append(movie)
                pass
                

            # movies = get_movies()
            # print(movies)

            print('movies:' + str(len(movie_hrefs)))

            excel = openpyxl.Workbook()
            sheet = excel.active
            sheet.title = "Top Rated IMDB AllTime"
            sheet.append(['Release Year', 'Name', 'Rating'])
            #Bold first row
            font = Font(bold=True)
            for cell in sheet["1:1"]:
                cell.font = font


            for movie_href in movie_hrefs:
                year = movie_href.find('span', class_='secondaryInfo').text
                name = movie_href.find('td', class_='titleColumn').a.text
                # genre = movie.find(
                #     'div', class_='lister-item-content').find('span', class_='genre').text.strip()
                rating = movie_href.find(
                    'td', class_='ratingColumn imdbRating').strong.text
                print(year + ' ' + name + ' ' + ' ' + rating)

                sheet.append([year, name, rating])

            # Insert top and bottom orders in each cell

            thin_border = Border(top=Side(style='thin'),
                                  bottom=Side(style='thin'))
            for cell in sheet["2:"+str(len(movie_hrefs)+1)]:
                for singlecell in cell:
                    print(singlecell)
                    singlecell.border = thin_border
            count = 0
            for cell in sheet["2:"+str(len(movie_hrefs)+1)]:
                for singlecell in cell:
                    if count % 2 == 0:
                        singlecell.fill = PatternFill(
                            "solid", start_color="AEAEAE")
                    else:
                        pass
                    print(singlecell)
                    singlecell.border = thin_border
                count += 1

            excel.save("Top Rated IMDB AllTime.xlsx")

        except Exception as e:
            print(e)
    return None


get_IMDB_movie_list(linkList)
